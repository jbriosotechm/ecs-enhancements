import os, sys
import openpyxl
import Config
import logging, traceback
import SystemConfig
import ReportConfig
import datetime
import time
import RP3Lib
import interactiveReport,convertDocxToDoc,traceback


def findStringLocationInSheet(sheet_obj,maxRows,maxColumns,expectedString):
    """
        Handle condition : what if the field is not found, should terminate execution with FATAL
    """
    for currentRow in range(1, maxRows+1):
        for currentCol in range(1,maxColumns+1):
            cell_obj = sheet_obj.cell(row=currentRow, column=currentCol)
            try:
                if str(cell_obj.value).strip()==str(expectedString):
                    return (currentRow,currentCol)
            except Exception,e:
                print traceback.print_exc()

    print "Terminating Execution since Excel Template was tampered."
    sys.exit(-1)



def createResultsFolderInRootFolder():

    #make sure root folder exists
    if not os.path.exists(ReportConfig.rootFolder):
        os.makedirs(ReportConfig.rootFolder)

    #create subFolder For Execution
    customTimestamp=str(datetime.datetime.now()).replace(" ", "_").replace(":", "_").replace(".", "").replace("-", "_")
    folderToCreate=ReportConfig.rootFolder+"\\Result_"+customTimestamp
    if not os.path.exists(folderToCreate):
        os.makedirs(folderToCreate)

    return (folderToCreate,customTimestamp)


def createCurrentResultExcel():

    (resultFolder,customTimestamp)=createResultsFolderInRootFolder()
    wb=openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ReportConfig.outputExcelPath=os.getcwd()+"\\"+resultFolder+"\\Report_"+customTimestamp+'.xlsx'
    print "[ Results will be logged to : ]"+ReportConfig.outputExcelPath
    wb.save(filename=ReportConfig.outputExcelPath)
    return (resultFolder,customTimestamp)


def basicOperations():
    wb_obj = openpyxl.load_workbook(ReportConfig.outputExcelPath)
    sheet_obj = wb_obj.active
    print "Sheet title :",sheet_obj.title
    maxRows=sheet_obj.max_row
    maxColumns=sheet_obj.max_column

    return (wb_obj,sheet_obj,maxRows,maxColumns)


def createBasicTemplate():

    (wb_obj,sheet_obj,maxRows,maxColumns)=basicOperations()
    ReportConfig.sheetObj=sheet_obj
    ReportConfig.workBookObj=wb_obj

    #write Excel Headers
    ColNumber=0
    RowNumber=1
    for header in ReportConfig.ExcelHeaders:
        ColNumber+=1
        setCellValue(sheet_obj,RowNumber,ColNumber,header)

    wb_obj.save(filename=ReportConfig.outputExcelPath)


def getNextRow():
    #check 2nd col which field in empty
    row=0
    colNumber=ReportConfig.ExcelHeaders.index(ReportConfig.TestStepNo)+1
    colNumber2=ReportConfig.ExcelHeaders.index(ReportConfig.TestScenarioSrNo)+1
    while True:
        row+=1

        if getCellValue(ReportConfig.sheetObj,row,colNumber) is None and getCellValue(ReportConfig.sheetObj,row,colNumber2) is None:
            break

    return row

def columnNameNumberMapping():
    for item in ReportConfig.ExcelHeaders:
        ReportConfig.columnNameToNumber[item]=ReportConfig.ExcelHeaders.index(item)+1
        print "ReportConfig.columnNameToNumber["+item+"]:",ReportConfig.columnNameToNumber[item]



def WriteTestCase(TestScenarioSrNo,TestCaseDesc):

    row=getNextRow()
    ReportConfig.currentTestCaseRowNumber=row
    startTime=str(datetime.datetime.now())+"\n[ "+str(time.time())+" ] "

    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.ExpectedResult], "ExpectedResult")
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.ActualResult],"ActualResult")
    setCellValue(ReportConfig.sheetObj,row,ReportConfig.columnNameToNumber[ReportConfig.TestScenarioSrNo],TestScenarioSrNo)
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.TestCaseStepDesc],TestCaseDesc)
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.StartTime],startTime)
    ReportConfig.workBookObj.save(filename=ReportConfig.outputExcelPath)



def WriteTestStep(TestStepDesc,ExpectedResult, ActualResult,StepStatus):
    row = getNextRow()
    ReportConfig.currentTestStepNumber+=1

    setCellValue(ReportConfig.sheetObj,row,ReportConfig.columnNameToNumber[ReportConfig.TestStepNo],str(ReportConfig.currentTestStepNumber))
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.TestCaseStepDesc],TestStepDesc)
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.ExpectedResult],ExpectedResult)
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.ActualResult],ActualResult)
    setCellValue(ReportConfig.sheetObj, row, ReportConfig.columnNameToNumber[ReportConfig.Status], StepStatus)
    ReportConfig.workBookObj.save(filename=ReportConfig.outputExcelPath)

def evaluateIfTestCaseIsPassOrFail():

    ReportConfig.currentTestStepNumber=0
    endTime = str(datetime.datetime.now()) + "\n[ " + str(time.time()) + " ] "
    status="Passed"
    for currentRow in range(ReportConfig.currentTestCaseRowNumber+1,getNextRow()):
        if "pass" not in getCellValue(ReportConfig.sheetObj,currentRow,ReportConfig.columnNameToNumber[ReportConfig.Status]).lower():
            status="Failed"
            break

    setCellValue(ReportConfig.sheetObj, ReportConfig.currentTestCaseRowNumber, ReportConfig.columnNameToNumber[ReportConfig.Status],status)
    setCellValue(ReportConfig.sheetObj,  ReportConfig.currentTestCaseRowNumber, ReportConfig.columnNameToNumber[ReportConfig.EndTime],endTime)
    ReportConfig.workBookObj.save(filename=ReportConfig.outputExcelPath)




def InitializeReporting():
    columnNameNumberMapping()
    #(folderToCreate,customTimestamp)=createResultsFolderInRootFolder()
    (folderToCreate,customTimestamp)=createCurrentResultExcel()
    createBasicTemplate()
    return folderToCreate






def getDelimiterFromExcel(sheet_obj,maxRows,maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.DelimiterWhileFileCreation)
    delim= (sheet_obj.cell(row=row, column=col+1)).value
    try:
        delim=delim.replace("[","").replace("]","")
        return delim
    except Exception,e:
        print "Unable to fetch Delimiter from Excel : Terminating Execution "
        traceback.print_exc()
        sys.exit(-1)


def addPadding(delimiter,stringToPad, columnWidth):
    return stringToPad.ljust(columnWidth,delimiter)


def createFile(sheet_obj, maxRows, maxColumns,dictColumnWidth,fileName,delimiter):
    #logic : get the row, col from where you need to start
    #col# will give you the dict count

    (row, col) = getStartingRowAndColumnForFileCreationTags(sheet_obj, maxRows, maxColumns)
    row+=1

    file=open(fileName,"w")

    # field len for each field should be stored in a dict, for faster processing
    for currentRow in range(row,maxRows+1):
        currentStr=""
        for currentCol in range(col,maxColumns+1):
            currentStr+=addPadding(delimiter,getCellValue(sheet_obj,currentRow,currentCol),int(dictColumnWidth[currentCol]))

        file.write(currentStr+"\n")


    file.close()


def getFileNameFromExcel(sheet_obj, maxRows, maxColumns):
    """
        The next column in the matched row will be returned
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.Souce_File_Name)
    fileName= (sheet_obj.cell(row=row, column=col+1)).value.strip()
    print "Source File Name : ",fileName
    return fileName

def getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns):
    """
    Returns row# and col# for the first tag of the file creation template
    In the same row, the cols shall continue till max col, take a note for capturing the last col
    """
    (row, col) = findStringLocationInSheet(sheet_obj, maxRows, maxColumns, SystemConfig.lastColumnBeforeFileCreationFields)
    return (row,col+1)




def getCellValue(sheet_obj,row,col):
    return sheet_obj.cell(row=row, column=col).value


def setCellValue(sheet_obj,row,col,valueToSet):
    try:
        sheet_obj.cell(row=row, column=col).value=valueToSet
    except:
        sheet_obj.cell(row=row, column=col).value="[INFO] Unable to render value since it contains characters which can't be rendered. Please check logs/console for the message"

def parseFieldTagsToSeperateNameAndWidth(cellValue):
    try:
        #print "Handling Parsing for Cell Value : ",cellValue
        fieldName=cellValue.split("[")[0].strip()
        fieldWidth=(cellValue.split("[")[1]).split("]")[0].strip()
        return (fieldName, fieldWidth)
    except Exception,e:
        print "Parsing Error : Terminating Execution since Excel Template was tampered / not created properly. Check the cellValue :",cellValue
        sys.exit(-1)




def storeFieldLengthsforAllFieldTags(sheet_obj,maxRows,maxColumns):
    dictColumnWidth={}
    (row,col)=getStartingRowAndColumnForFileCreationTags(sheet_obj,maxRows,maxColumns)
    #print "getStartingRowAndColumnForFileCreationTags() : Row : {0},Col : {1}".format(row,col)
    for currentCol in range(col,maxColumns+1):
        currentCellVal=getCellValue(sheet_obj,row,currentCol)
        (colName,colWidth)=parseFieldTagsToSeperateNameAndWidth(currentCellVal)
        dictColumnWidth[currentCol]=colWidth

    return dictColumnWidth


def printDict(dict):
    print "Printing Dictionary contents : ",dict


def GeneratePDFReport():
    scriptPath="NA"
    #outputExcelPath=ReportConfig.outputExcelPath
    #print "Output excel path : ",outputExcelPath

    #ReportConfig.outputExcelPath=r"C:\Users\aahmad\Desktop\Uni\PythonScripts\Results\Result_2020_01_02_14_59_26950000\\Report_2020_01_02_14_59_26950000.xlsx"
    outputExcelPath=ReportConfig.outputExcelPath
    outputPDFpath=str(ReportConfig.outputExcelPath).replace(".xlsx",".pdf")
    print("Output excel path: {0}".format(outputExcelPath))
    try:
        RP3Lib.main(scriptPath,outputExcelPath,outputExcelPath,outputPDFpath,"Report")
    except:
        print("[ERR] Issue generating word doc")
        traceback.print_exc()


    #convert docx to doc
    try:
        print("Output Excel path: ",outputExcelPath)
        convertDocxToDoc.main(outputExcelPath)
    except:
        print("[ERR] Issue generating PDF doc")
        traceback.print_exc()


    try:
        interactiveReport.main(outputExcelPath)
    except:
        print("[ERR] Issue generating word doc")
        traceback.print_exc()