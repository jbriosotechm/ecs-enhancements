import sys
import openpyxl

import SystemConfig
import userConfig

def read_sheet(sheet_name, last_column_name):
    wb_obj = openpyxl.load_workbook(userConfig.excelFileName)
    SystemConfig.sheetObj = wb_obj[sheet_name]
    SystemConfig.maxRows = SystemConfig.sheetObj.max_row + 1
    set_max_column(last_column_name)

def get_cell_location_of_string(expected_string):
    for current_row in range(1, SystemConfig.maxRows + 1):
        for current_col in range(1, SystemConfig.maxCol + 1):
            cell_value = get_cell_value(current_row, current_col)
            cell_value = str(cell_value).strip()
            if cell_value == str(expected_string):
                return [current_row, current_col]

    print "Expected String : {0} was not found in excel.".format(expected_string)
    print "Terminating Execution since Excel Template was tampered."
    sys.exit(-1)

def get_column_number_of_string(expected_string):
    return get_cell_location_of_string(expected_string)[1]

def get_row_number_of_string(expected_string):
    return get_cell_location_of_string(expected_string)[0]

def set_max_column(column_name):
    SystemConfig.maxCol = 100
    SystemConfig.maxCol = get_column_number_of_string(column_name)

def get_max_column():
    return SystemConfig.maxCol

def get_cell_value(row, col):
    return SystemConfig.sheetObj.cell(row, col).value

def read_results(filename):
    try:
        wb_obj = openpyxl.load_workbook(filename)
        sheetObj = wb_obj.active
    except Exception as e:
        return -1

    test_steps = []
    for row in sheetObj.iter_rows(min_row=3):
        step  = {}
        step['Step Description'] = row[2].value
        step['Expected'] = row[5].value
        step['Actual'] = row[6].value
        step['Status'] = row[7].value
        step['Screenshot'] = row[8].value
        test_steps.append(step)

    if len(test_steps) == 0:
        print ("[WARN] No Test Steps with status are Recorded in {0}".format(filename))
        return -2
    print(test_steps)
    return test_steps